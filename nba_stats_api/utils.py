import requests
import re
import time
import json
from decimal import Decimal
from bs4 import BeautifulSoup
from openpyxl import Workbook
from nba_py.player import PlayerShootingSplits, PlayerList
from nba_py.league import PlayerStats
from nba_py.constants import PerMode
from nba_stats_api.playtypes import PlayTypeHandler
from nba_stats_api.constants import (BasketballReference, ALL_PLAY_TYPES,
                                     COMMON_FIELDS, PLAYTYPE_COLUMNS,
                                     VIDEO_MEASURES, VIDEO_ENDPOINT,
                                     GENERAL_STAT_TYPES)
comm = re.compile("<!--|-->")


class DecimalEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, Decimal):
            return float(o)
        return super(DecimalEncoder, self).default(o)


class Player:
    def __init__(self, nba_id, name):
        self.nba_id = nba_id
        self.name = name


def get_shooting_stats(player_id, season):
    shooting_splits = PlayerShootingSplits(season=season,
                                           player_id=player_id)
    return shooting_splits.shot_5ft()[0] if len(shooting_splits.shot_5ft()) else {}


def extract_common_info(stats_dict):
    return {key: stats_dict[key] for key in COMMON_FIELDS}


def convert_dict(to_convert):
    converted = {key: to_convert[key] for key in to_convert if ("rank" not in key.lower()
                                                                and key not in COMMON_FIELDS
                                                                and key not in ["CFID", "CFPARAMS"])}
    return converted


def calc_true_shooting(player_stats_dict):
    totals = player_stats_dict['Totals']
    true_shooting_attempts = totals['FGA'] + (0.44 * totals['FTA'])
    if true_shooting_attempts != 0:
        true_shooting_percentage = Decimal(totals['PTS'] / (2 * true_shooting_attempts))
        return true_shooting_percentage
    return 0


def calc_three_point_attempt_rate(player_stats_dict):
    totals = player_stats_dict['Totals']
    fga = totals['FGA']
    if fga != 0:
        return Decimal(totals['FG3A'] / fga)
    else:
        return 0


def calc_free_throw_attempt_rate(player_stats_dict):
    totals = player_stats_dict['Totals']
    fga = totals['FGA']
    if fga != 0:
        return Decimal(totals['FTA'] / fga)
    else:
        return 0


def calc_two_pt_percentage(player_stats_dict):
    totals = player_stats_dict['Totals']
    two_pt_makes = totals['FGM'] - totals['FG3M']
    two_pt_attempts = totals['FGA'] - totals['FG3A']
    if two_pt_attempts != 0:
        return Decimal(two_pt_makes / two_pt_attempts)
    else:
        return 0


def calc_extra_shooting_stats(player_stats_dict):
    player_stats_dict['Shooting']['TS_PCT'] = calc_true_shooting(player_stats_dict)
    player_stats_dict['Shooting']['3PAr'] = calc_three_point_attempt_rate(player_stats_dict)
    player_stats_dict['Shooting']['FTr'] = calc_free_throw_attempt_rate(player_stats_dict)
    player_stats_dict['Shooting']['2PT_PCT'] = calc_two_pt_percentage(player_stats_dict)


def get_advanced_stats(bbref_id, season="2016-17"):
    url = BasketballReference.BaseURL + BasketballReference.PlayersEndpoint + bbref_id[0] + f"/{bbref_id}.html"
    response = requests.get(url)
    html = response.text
    cleaned_soup = BeautifulSoup(re.sub("<!--|-->", "", html), "html5lib")
    advanced_table = cleaned_soup.find("table", {'id': "advanced"})
    if advanced_table is not None:
        body = advanced_table.find("tbody")
        rows = body.find_all("tr")

        exclude_keys = ["yyy", "xxx", "age", "team_id", "lg_id", "pos", "g"]
        for row in rows:
            season_cell = row.find(attrs={'data-stat': "season"})
            if season_cell.text == season:
                cells = row.find_all("td")
                stats_dict = {tag.attrs['data-stat']: tag.text for tag in cells if
                              tag.attrs['data-stat'].lower() not in exclude_keys}
                return stats_dict
    else:
        return {}


def update_all_player_stats(season):
    player_stats_dict = {}
    pids_file = open("completed_pids.txt", "r")
    completed_pids = json.loads(pids_file.read())
    # completed_pids = [203518, 203112, 203500, 201167, 201582, 202332, 200746, 202730, 2754, 202329, 2365, 101187, 1626147, 203937, 201583, 203507, 2546, 201202, 2772, 201589, 201600, 201571, 202337, 1627758, 1627735, 2571, 200826, 203084, 2440, 203115, 101138, 201587, 201573, 203382, 203145, 203078, 1627736, 201563, 201158, 1627761, 1627733, 203461, 1627791, 202722, 201976, 202687, 202357, 204028, 202339, 202711, 101106, 1627762, 1626164, 202344, 202340, 201147, 1627763, 201166, 1626148, 201628, 1627759, 1627852, 203493, 203504, 202692, 202710, 203998, 101181, 203484, 203477, 203991, 201960, 1713, 203487, 201956, 1626161, 2199, 201163, 1627737, 1626176, 203902, 203546, 203903, 202708, 201954, 2555, 201144, 1626192, 1626188, 202326, 203496, 203459, 2037, 202348, 203109, 201967, 203552, 201939, 203584, 203076, 1627738, 202334, 201942, 203473, 1626155, 1627098, 203521, 2736, 1627767, 2564, 203476, 203915, 201962, 201609, 203083, 201162, 2399, 1627739, 201142, 204067, 1627740, 201961, 101145, 203954, 203898, 203516, 201936, 203957, 202702, 200770, 202324, 1627770, 1626245, 101109, 1627812, 1627827, 1627854, 203095, 200751, 204025, 101112, 201568, 204038, 1627868, 201188, 2200, 200752, 1627771, 202087, 202331, 1627875, 1626780, 201959, 1938, 203497, 203462, 203932, 201569, 101162, 1626203, 203924, 1626170, 201980, 203110, 101123, 203210, 201145, 201933, 203120, 1627773, 203501, 201935, 203090]
    for per_mode in [PerMode.Totals, PerMode.PerGame,
                     PerMode.Per100Possessions, PerMode.Per36]:
        print(f"Working on {per_mode}")
        player_stats = PlayerStats(season=season,
                                   per_mode=per_mode)
        time.sleep(1)
        for row in player_stats.overall():
            player_id = row['PLAYER_ID']
            new_row = convert_dict(row)
            if player_id in completed_pids:
                continue
            if player_id not in player_stats_dict:
                player_stats_dict[player_id] = {'BasicInfo': extract_common_info(row)}

            player_stats_dict[player_id][per_mode] = new_row

    play_type_handler = PlayTypeHandler(year=season.split("-")[0],
                                        season_type="Reg")
    for play_type in ALL_PLAY_TYPES:
        print(f"Working on {play_type}")
        play_type_handler.fetch_json(play_type)

        for player_id in play_type_handler.json:
            if player_id in completed_pids:
                continue
            player_name = play_type_handler.json[player_id]['PLAYER_NAME']
            if player_id not in player_stats_dict:
                print(f"{player_name} has entries for PlayType Statistics, but not"
                      f"traditional stats. Skipping this player.")
                continue
            player_stats_dict[player_id][play_type] = convert_dict(play_type_handler.json[player_id])

        time.sleep(1)

    player_list = PlayerList(season=season,
                             only_current=1)

    try:
        with open("bbref_id_map.json", "r") as bbref_file:
            bbref_id_map = json.loads(bbref_file.read())
            for player in player_list.info():
                player_id = player['PERSON_ID']
                if player_id in completed_pids:
                    continue
                player_name = player['DISPLAY_FIRST_LAST']
                print(f"Working on shooting for {player_name}")
                shooting = get_shooting_stats(player_id=player_id,
                                              season=season)
                if player_id in player_stats_dict:
                    player_stats_dict[player_id]['Shooting'] = convert_dict(shooting)
                    calc_extra_shooting_stats(player_stats_dict[player_id])

                    bbref_id = bbref_id_map[str(player_id)]
                    print(f"Working on advanced for {player_name}")
                    advanced_stats = get_advanced_stats(bbref_id, season="2016-17")
                    player_stats_dict[player_id]['Advanced'] = advanced_stats
                    completed_pids.append(player_id)
                else:
                    print(f"It appears {player_name} hasn't registered any statistics for this season.")

                time.sleep(1)
    except Exception as e:
        pids_file = open("completed_pids.txt", "w")
        pids_file.write(json.dumps(completed_pids))
        pids_file.close()
        print(("Num completed", len(completed_pids)))
        raise e
    return player_stats_dict


def generate_excel_spreadsheet(player_stats, season):
    wkbook = Workbook()
    worksheet = wkbook.active
    worksheet['A1'] = player_stats['BasicInfo']['PLAYER_NAME'] + f"{season} stats and video hub"
    worksheet['A4'] = "Totals"

    cur_col = 1
    cur_row = 4
    for stat_type in GENERAL_STAT_TYPES:
        worksheet.cell(row=cur_row, column=cur_col, value=stat_type)
        cur_row += 1
        for column in player_stats[stat_type]:
            worksheet.cell(row=cur_row, column=cur_col, value=column)

            value = player_stats[stat_type][column]
            worksheet.cell(row=cur_row + 1, column=cur_col, value=value)
            cur_col += 1
        cur_col = 1
        cur_row += 3

    cur_row += 2
    worksheet.cell(row=cur_row, column=cur_col, value="Play Type")
    cur_col += 1
    for column in PLAYTYPE_COLUMNS:
        worksheet.cell(row=cur_row, column=cur_col, value=column)
        cur_col += 1
    cur_row += 1
    cur_col = 1
    video_col = cur_col
    video_row = cur_row

    for play_type in ALL_PLAY_TYPES:
        worksheet.cell(row=cur_row, column=cur_col, value=play_type)
        cur_col += 1
        for column in player_stats.get(play_type, []):
            worksheet.cell(row=cur_row, column=cur_col, value=player_stats[play_type][column])
            cur_col += 1
            video_col = cur_col
        cur_col = 1
        cur_row += 1

    video_col += 1
    worksheet.cell(row=video_row, column=video_col, value="Video (NBA.com archive)")

    for measure in VIDEO_MEASURES:
        url = VIDEO_ENDPOINT.format(player_id=player_stats['BasicInfo']['PLAYER_ID'],
                                    measure=measure, season=season,
                                    season_type="Regular+Season",
                                    )
        worksheet.cell(row=video_row, column=video_col, value=VIDEO_MEASURES[measure])
        worksheet.cell(row=video_row, column=video_col).hyperlink = url
        video_row += 1

    player_name = player_stats['BasicInfo']['PLAYER_NAME']
    wkbook.save(f"outputs/{player_name}.xlsx")
